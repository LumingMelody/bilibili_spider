import random
import json
import time

import requests
from openpyxl import Workbook

USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_8; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50",
    "Mozilla/5.0 (Windows; U; Windows NT 6.1; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50",
    "Mozilla/5.0 (Windows NT 10.0; WOW64; rv:38.0) Gecko/20100101 Firefox/38.0",
    "Mozilla/5.0 (Windows NT 10.0; WOW64; Trident/7.0; .NET4.0C; .NET4.0E; .NET CLR 2.0.50727; .NET CLR 3.0.30729; .NET CLR 3.5.30729; InfoPath.3; rv:11.0) like Gecko",
    "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0)",
    "Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0)",
    "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1)",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.6; rv:2.0.1) Gecko/20100101 Firefox/4.0.1",
    "Mozilla/5.0 (Windows NT 6.1; rv:2.0.1) Gecko/20100101 Firefox/4.0.1",
    "Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; en) Presto/2.8.131 Version/11.11",
    "Opera/9.80 (Windows NT 6.1; U; en) Presto/2.8.131 Version/11.11",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_0) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Maxthon 2.0)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; TencentTraveler 4.0)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; The World)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; SE 2.X MetaSr 1.0; SE 2.X MetaSr 1.0; .NET CLR 2.0.50727; SE 2.X MetaSr 1.0)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; 360SE)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Avant Browser)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1)",
    "Mozilla/5.0 (iPhone; U; CPU iPhone OS 4_3_3 like Mac OS X; en-us) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8J2 Safari/6533.18.5",
    "Mozilla/5.0 (iPod; U; CPU iPhone OS 4_3_3 like Mac OS X; en-us) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8J2 Safari/6533.18.5",
    "Mozilla/5.0 (iPad; U; CPU OS 4_3_3 like Mac OS X; en-us) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8J2 Safari/6533.18.5",
    "Mozilla/5.0 (Linux; U; Android 2.3.7; en-us; Nexus One Build/FRF91) AppleWebKit/533.1 (KHTML, like Gecko) Version/4.0 Mobile Safari/533.1",
    "MQQBrowser/26 Mozilla/5.0 (Linux; U; Android 2.3.7; zh-cn; MB200 Build/GRJ22; CyanogenMod-7) AppleWebKit/533.1 (KHTML, like Gecko) Version/4.0 Mobile Safari/533.1",
    "Opera/9.80 (Android 2.3.4; Linux; Opera Mobi/build-1107180945; U; en-GB) Presto/2.8.149 Version/11.10",
    "Mozilla/5.0 (Linux; U; Android 3.0; en-us; Xoom Build/HRI39) AppleWebKit/534.13 (KHTML, like Gecko) Version/4.0 Safari/534.13",
    "Mozilla/5.0 (BlackBerry; U; BlackBerry 9800; en) AppleWebKit/534.1+ (KHTML, like Gecko) Version/6.0.0.337 Mobile Safari/534.1+",
    "Mozilla/5.0 (hp-tablet; Linux; hpwOS/3.0.0; U; en-US) AppleWebKit/534.6 (KHTML, like Gecko) wOSBrowser/233.70 Safari/534.6 TouchPad/1.0",
    "Mozilla/5.0 (SymbianOS/9.4; Series60/5.0 NokiaN97-1/20.0.019; Profile/MIDP-2.1 Configuration/CLDC-1.1) AppleWebKit/525 (KHTML, like Gecko) BrowserNG/7.1.18124",
    "Mozilla/5.0 (compatible; MSIE 9.0; Windows Phone OS 7.5; Trident/5.0; IEMobile/9.0; HTC; Titan)",
    "UCWEB7.0.2.37/28/999",
    "NOKIA5700/ UCWEB7.0.2.37/28/999",
    "Openwave/ UCWEB7.0.2.37/28/999",
    "Mozilla/4.0 (compatible; MSIE 6.0; ) Opera/UCWEB7.0.2.37/28/999",
]

headers = {
    'User-Agent': random.choice(USER_AGENTS),
    'cookie': "_uuid=19E3D158-5C4C-5313-15F4-D0E0652B856B61295infoc; buvid3=BA6017C5-E479-4596-97FA-EF03FC1E1435185012infoc; buvid_fp=BA6017C5-E479-4596-97FA-EF03FC1E1435185012infoc; LIVE_BUVID=AUTO8616167213810503; fingerprint=3057ac8ee6ac336ca85353bf3ade1252; buvid_fp_plain=7E65C238-AD53-4548-A5A5-AB7C47EE239D185015infoc; SESSDATA=a050ad08%2C1632555146%2C8004d%2A31; bili_jct=b82a7d19334d1f1cbbc1d74982f871e4; DedeUserID=20221702; DedeUserID__ckMd5=d6786b5d622ed8dc; sid=dbdf5h6y; CURRENT_FNVAL=80; blackside_state=1; bsource=search_baidu; PVID=1"
}
wb = Workbook()
wb1 = Workbook()
wb2 = Workbook()

ws = wb.active
ws1 = wb1.active
ws2 = wb2.active
# ws1 = wb.create_sheet()

ws.append(
    [
        "投稿标题",
        "投稿描述",
        "动态内容",
        "转发数"
        "点赞数",
        "评论数",
        "发布时间",
    ]
)

ws1.append([
    "转发内容",
    "点赞数",
    "转发数",
    "评论数",
    "转发时间",
])

ws2.append([
    "动态内容",
    "点赞数",
    "转发数",
    "评论数",
    "上传时间",
])


def get_bilibli_dynamic(b_url):
    resp = requests.get(url=b_url, headers=headers).json()
    # print(resp)
    while resp['data']['next_offset'] is not None:
        cards = resp['data']['cards']
        for card in cards:
            content = json.loads(card['card'])
            print(content)
            if 'aid' in content.keys():
                likes = card['desc']['like']
                repost = card['desc']['repost']
                dynamic = content['dynamic']
                desc = content['desc']
                title = content['title']
                reply = content['stat']['reply']
                ts = content['pubdate']
                timeArray = time.localtime(ts)
                pubdate = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
                ws.append([title, desc, dynamic, repost, likes, reply, pubdate])
                wb.save(r"D:\bilibili\bilibli_07_21\bili_07_21_五菱汽车_投稿.xlsx")
            elif 'origin_user' in content.keys():
                contents = content['item']['content']
                likes = card['desc']['like']
                reposts = card['desc']['repost']
                comments = card['desc']['comment']
                ts = content['item']['timestamp']
                timeArray = time.localtime(ts)
                create_time = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
                # item_data = json.loads(content['item']['origin'])
                # desc = item_data['desc']
                ws1.append([contents, likes, reposts, comments, create_time])
                wb1.save(r"D:\bilibili\bilibli_07_21\bili_07_21_五菱汽车_转发.xlsx")
            elif "description" in content['item'].keys():
                likes = card['desc']['like']
                reposts = card['desc']['repost']
                comments = card['desc']['comment']
                desc = content['item']['description']
                ts = content['item']['upload_time']
                timeArray = time.localtime(ts)
                upload_time = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
                ws2.append([desc, likes, reposts, comments, upload_time])
                wb2.save(r"D:\bilibili\bilibli_07_21\bili_07_21_五菱汽车_动态.xlsx")
        next_offset = resp['data']['next_offset']
        next_url = f"https://api.vc.bilibili.com/dynamic_svr/v1/dynamic_svr/space_history?visitor_uid=20221702&host_uid=439868658&offset_dynamic_id={next_offset}&need_top=1&platform=web"
        resp = get_bilibli_dynamic(next_url)
    # wb.save(r"D:\bilibili\bilibli_07_21\bili_07_21_五菱汽车_动态.xlsx")


if __name__ == '__main__':
    url = "https://api.vc.bilibili.com/dynamic_svr/v1/dynamic_svr/space_history?visitor_uid=20221702&host_uid=439868658&offset_dynamic_id=0&need_top=1&platform=web"
    get_bilibli_dynamic(url)
